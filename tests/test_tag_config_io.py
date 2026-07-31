from io import BytesIO
import zipfile

from openpyxl import load_workbook
import pytest

from pca_model_builder.tag_config_io import (
    TAG_CONFIG_HEADERS,
    build_tag_config_template,
    export_tag_config_workbook,
    parse_tag_config_workbook,
)


def test_template_has_fixed_sheet_headers_prefilled_tags_and_role_validation():
    content = build_tag_config_template(["TI001", "PI001"])
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Tags"]

    assert workbook.sheetnames == ["Tags"]
    assert tuple(cell.value for cell in sheet[1]) == TAG_CONFIG_HEADERS
    assert [sheet["A2"].value, sheet["A3"].value] == ["TI001", "PI001"]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:K3"
    assert sheet.data_validations.count == 1


def test_import_allows_blank_fields_and_reports_unknown_duplicate_and_invalid_rows():
    content = export_tag_config_workbook(
        ["TI001", "PI001"],
        {
            "TI001": {"description": "温度"},
            "PI001": {"role": "state_filter"},
        },
    )
    parsed = parse_tag_config_workbook(content, ["TI001", "PI001"])

    assert parsed["can_apply"]
    assert parsed["matched_tags"] == ["TI001", "PI001"]
    assert parsed["configs"]["TI001"]["description"] == "温度"
    assert parsed["configs"]["PI001"]["role"] == "state_filter"

    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Tags"]
    sheet.append(["TI001"])
    sheet.append(["UNKNOWN"])
    sheet.append(["BAD", None, None, "invalid_role"])
    broken = BytesIO()
    workbook.save(broken)
    preview = parse_tag_config_workbook(broken.getvalue(), ["TI001", "PI001"])

    assert not preview["can_apply"]
    assert preview["duplicate_tags"] == ["TI001"]
    assert preview["unknown_template_tags"] == ["BAD", "UNKNOWN"]
    assert any("role非法" in error for error in preview["errors"])


def test_export_reimports_consistently_and_invalid_files_are_rejected():
    configs = {
        "TI001": {
            "description": "温度",
            "unit": "℃",
            "engineering_min": 0,
            "engineering_max": 300,
            "comment": "反应器",
        }
    }
    content = export_tag_config_workbook(["TI001"], configs)
    parsed = parse_tag_config_workbook(content, ["TI001"])

    assert parsed["configs"]["TI001"]["engineering_max"] == 300.0
    assert parsed["configs"]["TI001"]["comment"] == "反应器"
    with pytest.raises(ValueError, match="为空"):
        parse_tag_config_workbook(b"", ["TI001"])
    with pytest.raises(ValueError, match="无法读取"):
        parse_tag_config_workbook(b"not an xlsx", ["TI001"])


@pytest.mark.parametrize(
    ("value", "message"),
    [("not-a-number", "must be numeric"), (20, "lower value must be less")],
)
def test_import_preview_reports_invalid_numeric_and_range(value, message):
    content = build_tag_config_template(["TI001"])
    workbook = load_workbook(BytesIO(content))
    sheet = workbook["Tags"]
    sheet["E2"] = value
    sheet["F2"] = 10
    broken = BytesIO()
    workbook.save(broken)

    preview = parse_tag_config_workbook(broken.getvalue(), ["TI001"])

    assert not preview["can_apply"]
    assert any(message in error for error in preview["errors"])


def test_import_rejects_workbook_containing_macro_payload():
    content = build_tag_config_template(["TI001"])
    macro = BytesIO()
    with zipfile.ZipFile(BytesIO(content)) as source:
        with zipfile.ZipFile(macro, "w", zipfile.ZIP_DEFLATED) as destination:
            for name in source.namelist():
                destination.writestr(name, source.read(name))
            destination.writestr("xl/vbaProject.bin", b"macro")

    with pytest.raises(ValueError, match="包含宏"):
        parse_tag_config_workbook(macro.getvalue(), ["TI001"])
