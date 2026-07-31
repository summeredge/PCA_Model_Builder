from __future__ import annotations

from io import BytesIO
from typing import Any, Mapping, Sequence
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

from .tag_config import TAG_ROLES, normalize_tag_registry


TAG_CONFIG_HEADERS = (
    "tag",
    "description",
    "unit",
    "role",
    "engineering_min",
    "engineering_max",
    "normal_min",
    "normal_max",
    "alarm_min",
    "alarm_max",
    "comment",
)
MAX_TAG_CONFIG_BYTES = 10 * 1024 * 1024
_HEADER_NOTES = {
    "tag": "必须与历史CSV列名完全一致。",
    "role": "continuous_input/state_filter/label_only/exclude",
    "engineering_min": "可选；填写时必须同时填写engineering_max。",
    "engineering_max": "可选；必须大于engineering_min。",
    "normal_min": "可选；仅用于工程解释。",
    "normal_max": "可选；仅用于工程解释。",
    "alarm_min": "可选；仅用于工程解释。",
    "alarm_max": "可选；仅用于工程解释。",
}


def build_tag_config_template(tags: Sequence[str]) -> bytes:
    registry = {tag: {} for tag in tags}
    return export_tag_config_workbook(tags, registry)


def export_tag_config_workbook(
    tags: Sequence[str], configs: Mapping[str, Mapping[str, Any]] | None
) -> bytes:
    normalized = normalize_tag_registry(tags, configs)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tags"
    sheet.append(TAG_CONFIG_HEADERS)
    for column, header in enumerate(TAG_CONFIG_HEADERS, 1):
        cell = sheet.cell(1, column)
        cell.comment = Comment(_HEADER_NOTES.get(header, "可选工程元数据。"), "Codex")
    for tag in tags:
        config = normalized[tag]
        sheet.append(
            [
                tag,
                config["description"],
                config["unit"],
                config["role"],
                config["engineering_min"],
                config["engineering_max"],
                config["normal_min"],
                config["normal_max"],
                config["alarm_min"],
                config["alarm_max"],
                config["comment"],
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:K{max(2, len(tags) + 1)}"
    validation = DataValidation(
        type="list",
        formula1='"continuous_input,state_filter,label_only,exclude"',
        allow_blank=False,
    )
    sheet.add_data_validation(validation)
    validation.add(f"D2:D{max(2, len(tags) + 1)}")
    widths = (24, 24, 12, 20, 16, 16, 14, 14, 14, 14, 28)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def parse_tag_config_workbook(
    content: bytes, data_tags: Sequence[str]
) -> dict[str, Any]:
    if not content:
        raise ValueError("XLSX配置文件为空")
    if len(content) > MAX_TAG_CONFIG_BYTES:
        raise ValueError("XLSX配置文件超过10 MB限制")
    try:
        with zipfile.ZipFile(BytesIO(content)) as package:
            if any(
                name.lower().endswith("vbaproject.bin")
                for name in package.namelist()
            ):
                raise ValueError("不接受包含宏的XLSX配置文件")
    except zipfile.BadZipFile as error:
        raise ValueError("无法读取XLSX配置文件") from error
    try:
        workbook = load_workbook(
            BytesIO(content), read_only=True, data_only=True, keep_vba=False
        )
    except Exception as error:
        raise ValueError("无法读取XLSX配置文件") from error
    if workbook.sheetnames != ["Tags"]:
        raise ValueError("XLSX配置必须且只能包含Tags工作表")
    sheet = workbook["Tags"]
    header = tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
    if header != TAG_CONFIG_HEADERS:
        raise ValueError("Tags工作表表头不符合模板")

    rows: dict[str, dict[str, Any]] = {}
    duplicates: list[str] = []
    errors: list[str] = []
    data_tag_set = set(data_tags)
    unknown: list[str] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in {None, ""} for value in values):
            continue
        tag = str(values[0] or "").strip()
        if not tag:
            errors.append("空白Tag：tag不能为空")
            continue
        if tag in rows:
            duplicates.append(tag)
            continue
        raw = {
            field: value
            for field, value in zip(TAG_CONFIG_HEADERS[1:], values[1:], strict=True)
            if value not in {None, ""}
        }
        role = str(raw.get("role", "continuous_input")).strip()
        if role not in TAG_ROLES:
            errors.append(f"{tag}：role非法（{role}）")
        rows[tag] = raw
        if tag not in data_tag_set:
            unknown.append(tag)
    for tag in sorted(set(duplicates)):
        errors.append(f"{tag}：模板中重复出现")
    for tag in sorted(set(unknown)):
        errors.append(f"{tag}：当前历史数据中不存在")

    matched = [tag for tag in data_tags if tag in rows]
    valid_configs: dict[str, dict[str, Any]] = {}
    for tag in matched:
        try:
            valid_configs[tag] = normalize_tag_registry([tag], {tag: rows[tag]})[tag]
        except ValueError as error:
            errors.append(str(error))
    return {
        "configs": valid_configs,
        "provided_configs": {
            tag: rows[tag] for tag in matched if tag in valid_configs
        },
        "matched_tags": matched,
        "unconfigured_data_tags": [tag for tag in data_tags if tag not in rows],
        "unknown_template_tags": sorted(set(unknown)),
        "duplicate_tags": sorted(set(duplicates)),
        "errors": errors,
        "can_apply": not errors,
    }
